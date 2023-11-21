from datetime import datetime, timedelta

from openpyxl import load_workbook

from sql.models import DataTypes, QTypes
from sql.schemas import TableBase


def convert_row_to_pydantic_object(row, date: datetime) -> list[TableBase]:
    id = row[0].value
    company_name = row[1].value
    result = []
    for i in range(2, 10):
        data_type = DataTypes.fact if i > 6 else DataTypes.forecast
        q_type = QTypes.qliq if i in [2, 3, 6, 7] else QTypes.qoil
        result.append(
            TableBase(
                entry_id=id,
                company=company_name,
                data_type=data_type,
                q_type=q_type,
                value=row[i].value,
                date=date,
            )
        )
    return result


def read_exel(file_path: str) -> list[TableBase] | None:
    try:
        wb = load_workbook(file_path)
        sheet = wb.active
        if sheet:
            result = []
            now = datetime.now()
            now_month = datetime.now() - timedelta(now.day)
            for i, row in enumerate(sheet.iter_rows()):
                if i < 3:
                    continue
                date = now_month + timedelta(days=i)
                converted_row = convert_row_to_pydantic_object(row, date)
                result.extend(converted_row)
            return result
        else:
            print("sheet is None")
    except Exception as ex:
        print(ex)
